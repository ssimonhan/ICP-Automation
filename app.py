import streamlit as st
import tempfile
import os
from pathlib import Path
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from calibration_settings import CalibrationRange, HighlightThresholds
from process_icp_data import normalize_sample_initials, process_icp_file
from generate_icp_concentrations import build_concentration_workbook
from excel_backend import excel_backend

st.set_page_config(page_title="ICP Automation", layout="wide")

st.title("🧪 ICP Automation Tool")

if excel_backend() == "cloud":
    st.info(
        "ℹ️ Running in cloud mode. "
        "Excel automatic formatting is approximated for compatibility."
    )

uploaded_file = st.file_uploader("Upload ICP export", type=["csv", "xlsx"])
initials = st.text_input("Sample initials", value="SH")

with st.expander("Advanced settings"):
    col1, col2 = st.columns(2)
    with col1:
        st.caption("Customize ppb ranges used for ICP highlighting and final concentration selection.")
        orange_min = st.number_input(
            "Orange range minimum ppb",
            min_value=0.0,
            value=1.0,
            step=0.1,
            help="Values from this number up to the green minimum are lower-confidence and highlighted orange.",
        )
        green_min = st.number_input(
            "Green range minimum ppb",
            min_value=0.0,
            value=10.0,
            step=1.0,
            help="Values from this number through the green maximum are treated as accurate.",
        )
        green_max = st.number_input(
            "Green range maximum ppb",
            min_value=0.0,
            value=400.0,
            step=10.0,
            help="Values greater than this are outside the accurate range.",
        )
    with col2:
        st.caption("Customize internal-standard and blank-sample warning thresholds.")
        internal_light_orange = st.number_input(
            "Internal standard tolerance before orange (%)",
            min_value=0.0,
            value=20.0,
            step=1.0,
        )
        internal_orange = st.number_input(
            "Internal standard orange threshold (%)",
            min_value=0.0,
            value=40.0,
            step=1.0,
            help="Values above this threshold are highlighted dark orange.",
        )
        blank_light_orange = st.number_input(
            "Blank sample tolerance before orange (ppb)",
            min_value=0.0,
            value=0.2,
            step=0.1,
        )
        blank_orange = st.number_input(
            "Blank sample orange threshold (ppb)",
            min_value=0.0,
            value=1.0,
            step=0.1,
            help="Values above this threshold are highlighted dark orange.",
        )

try:
    sample_initials = normalize_sample_initials(initials)
    calibration_range = CalibrationRange(
        orange_min=orange_min,
        green_min=green_min,
        green_max=green_max,
    )
    highlight_thresholds = HighlightThresholds(
        internal_light_orange=internal_light_orange,
        internal_orange=internal_orange,
        blank_light_orange=blank_light_orange,
        blank_orange=blank_orange,
    )
except ValueError as exc:
    st.error(f"Settings error: {exc}")
    st.stop()

# Store intermediate file path
if "mid_path" not in st.session_state:
    st.session_state.mid_path = None

# =========================
# STEP 1
# =========================

if uploaded_file and st.button("Step 1: Generate ICP Sheet"):

    with st.spinner("Processing raw file..."):

        with tempfile.NamedTemporaryFile(delete=False, suffix=uploaded_file.name) as tmp_in:
            tmp_in.write(uploaded_file.getvalue())
            input_path = Path(tmp_in.name)

        tmp_mid = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        mid_path = Path(tmp_mid.name)
        tmp_mid.close()

        process_icp_file(
            input_path=input_path,
            output_path=mid_path,
            sheet_name=None,
            sample_initials=sample_initials,
            calibration_range=calibration_range,
            highlight_thresholds=highlight_thresholds,
        )

        os.remove(input_path)

    # Save path
    st.session_state.mid_path = str(mid_path)

    # Store file bytes persistently
    with open(mid_path, "rb") as f:
        st.session_state.cleaned_file = f.read()

    st.success("✅ ICP sheet generated. Now edit dilution and sample below.")
    
    st.session_state.pop("final_file", None)

# ALWAYS show download button if file exists
if "cleaned_file" in st.session_state:
    st.download_button(
        "⬇️ Download Cleaned ICP File",
        data=st.session_state.cleaned_file,
        file_name=f"{initials}_ICP_cleaned.xlsx"
    )

# =========================
# STEP 2 — EDIT TABLE
# =========================

if st.session_state.mid_path:

    wb = load_workbook(st.session_state.mid_path)
    
    # Find ICP sheet
    icp_sheet_name = None
    for name in wb.sheetnames:
        if name.upper().endswith(" ICP"):
            icp_sheet_name = name
            break

    if icp_sheet_name is None:
        st.error("No ICP sheet found")
        st.stop()

    ws = wb[icp_sheet_name]

    # Convert to DataFrame
    data = []

    for row in range(3, ws.max_row + 1):
        sample_name = ws.cell(row, 3).value

        if sample_name:
            data.append({
                "Row": row,
                "Dilution": ws.cell(row, 1).value,
                "Sample": ws.cell(row, 2).value or "",
                "Sample Name": sample_name
            })

    df = pd.DataFrame(data)

    st.subheader("✏️ Edit Dilution and Sample")

    edited_df = st.data_editor(
        df,
        width='stretch',
        disabled=["Row", "Sample Name"],  # prevent breaking keys        
        column_config={
            "Dilution": st.column_config.NumberColumn("Dilution", step=1)
        }
    )

    # =========================
    # STEP 3 — RUN FINAL
    # =========================

    if st.button("Step 2: Generate Final Concentrations"):

        if edited_df["Dilution"].isnull().any():
            st.error("❌ Please fill in all dilution values.")
            st.stop()

        with st.spinner("Running full analysis..."):

            # Write edits back to workbook
            wb = load_workbook(st.session_state.mid_path)
            ws = wb[icp_sheet_name]

            for _, row in edited_df.iterrows():
                excel_row = int(row["Row"])
                ws.cell(excel_row, 1).value = row["Dilution"]
                ws.cell(excel_row, 2).value = row["Sample"]

            wb.save(st.session_state.mid_path)

            # Output file
            tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            output_path = Path(tmp_out.name)
            tmp_out.close()

            build_concentration_workbook(
                source_path=Path(st.session_state.mid_path),
                output_path=output_path,
                source_sheet_name=icp_sheet_name,
                calibration_range=calibration_range,
            )

        with open(output_path, "rb") as f:
            st.session_state.final_file = f.read()

        st.success("✅ Final concentrations generated!")

# Persistent download button
if "final_file" in st.session_state:
    st.download_button(
        "⬇️ Download Final Excel",
        data=st.session_state.final_file,
        file_name="icp_results.xlsx"
    )
