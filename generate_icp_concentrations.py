"""Generate an ICP concentration workbook from an edited <initials> ICP sheet."""

from __future__ import annotations

__author__ = "Shihua Han"
__version__ = "0.4.1"

import argparse
import re
import subprocess
import warnings
from collections import OrderedDict
from copy import copy
from pathlib import Path

from calibration_settings import DEFAULT_CALIBRATION_RANGE, CalibrationRange, format_number
from excel_backend import excel_backend
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, CellIsRule


DEFAULT_SOURCE_SHEET_NAME = "SH ICP"
OUTPUT_SHEET_NAME = "ICP concentrations"
HEADER_ROW = 1
DATA_START_ROW = 3
ELEMENT_START_COL = 4

OUTLINE_SIDE = Side(style="thin", color="808080")
FILL_GOLD = PatternFill("solid", fgColor="FFD966")
EXCEL_FILL_GREEN = 13561798
EXCEL_FILL_ORANGE = 14083324
EXCEL_FILL_GREY = 14277081
EXCEL_FILL_RED = 13421823
EXCEL_ICP_RANGE_FILL = 13693658

warnings.filterwarnings(
    "ignore",
    message="Unknown extension is not supported and will be removed",
    category=UserWarning,
)
warnings.filterwarnings(
    "ignore",
    message="Conditional Formatting extension is not supported and will be removed",
    category=UserWarning,
)


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text or text.upper() in {"N/A", "NA", "NULL", "NONE"}:
        return None
    if text.startswith("<"):
        text = text[1:].strip()

    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text)
    return float(match.group(0)) if match else None


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def copy_cell(source, target) -> None:
    target.value = source.value
    copy_cell_style(source, target)
    if source.comment:
        target.comment = copy(source.comment)


def apply_outline_border(output_ws, first_row: int, last_row: int, first_col: int, last_col: int) -> None:
    """Apply only the outside border around a rectangular section."""

    for row in range(first_row, last_row + 1):
        for col in range(first_col, last_col + 1):
            cell = output_ws.cell(row, col)
            existing = cell.border
            cell.border = Border(
                left=OUTLINE_SIDE if col == first_col else existing.left,
                right=OUTLINE_SIDE if col == last_col else existing.right,
                top=OUTLINE_SIDE if row == first_row else existing.top,
                bottom=OUTLINE_SIDE if row == last_row else existing.bottom,
                diagonal=existing.diagonal,
                diagonal_direction=existing.diagonal_direction,
                diagonalUp=existing.diagonalUp,
                diagonalDown=existing.diagonalDown,
                outline=existing.outline,
                vertical=existing.vertical,
                horizontal=existing.horizontal,
            )


def read_groups(source_ws) -> tuple[list[str], OrderedDict[str, list[dict[str, object]]]]:
    elements = [normalize(source_ws.cell(HEADER_ROW, col).value) for col in range(ELEMENT_START_COL, source_ws.max_column + 1)]
    groups: OrderedDict[str, list[dict[str, object]]] = OrderedDict()
    current_sample = ""

    for row_idx in range(DATA_START_ROW, source_ws.max_row + 1):
        sample_name = normalize(source_ws.cell(row_idx, 3).value)
        if not sample_name:
            continue

        sample = normalize(source_ws.cell(row_idx, 2).value) or current_sample
        if sample:
            current_sample = sample
        else:
            sample = "Unspecified sample"

        groups.setdefault(sample, []).append(
            {
                "source_row": row_idx,
                "dilution": parse_number(source_ws.cell(row_idx, 1).value),
                "sample": sample,
                "sample_name": sample_name,
                "ppb_values": [
                    parse_number(source_ws.cell(row_idx, col).value)
                    for col in range(ELEMENT_START_COL, source_ws.max_column + 1)
                ],
            }
        )

    return elements, groups


def selected_index(
    ppb_values: list[float | None],
    dilution_values: list[float | None],
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
) -> int | None:
    indexed = [
        (idx, value, dilution)
        for idx, (value, dilution) in enumerate(zip(ppb_values, dilution_values))
        if value is not None and dilution is not None
    ]
    if not indexed:
        return None

    green_values = [
        (idx, value, dilution)
        for idx, value, dilution in indexed
        if calibration_range.green_min <= value <= calibration_range.green_max
    ]
    if green_values:
        return min(green_values, key=lambda item: item[2])[0]

    orange_values = [
        (idx, value, dilution)
        for idx, value, dilution in indexed
        if calibration_range.orange_min <= value < calibration_range.green_min
    ]
    if orange_values:
        return min(orange_values, key=lambda item: item[2])[0]

    if all(value < calibration_range.orange_min for _, value, _ in indexed):
        return min(indexed, key=lambda item: item[2])[0]

    if all(value > calibration_range.green_max for _, value, _ in indexed):
        return max(indexed, key=lambda item: item[2])[0]

    return None


def write_header(output_ws, source_ws, output_row: int, elements: list[str]) -> None:
    for col in range(1, ELEMENT_START_COL):
        copy_cell(source_ws.cell(HEADER_ROW, col), output_ws.cell(output_row, col))
    for offset, element in enumerate(elements):
        col = ELEMENT_START_COL + offset
        copy_cell_style(source_ws.cell(HEADER_ROW, col), output_ws.cell(output_row, col))
        output_ws.cell(output_row, col).value = element


def write_sample_block(
    output_ws,
    source_ws,
    start_row: int,
    sample: str,
    rows: list[dict[str, object]],
    elements: list[str],
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
) -> tuple[int, list[int], tuple[int, int, int, int, int]]:
    title_row = start_row
    header_row = start_row + 1
    ppb_start = header_row + 1
    ppm_start = ppb_start + len(rows)
    selected_row = ppm_start + len(rows)
    last_col = ELEMENT_START_COL + len(elements) - 1
    data_bar_rows: list[int] = []

    output_ws.cell(title_row, 1, sample)
    output_ws.cell(title_row, 1).font = copy(source_ws.cell(HEADER_ROW, 1).font)

    write_header(output_ws, source_ws, header_row, elements)

    for offset, row in enumerate(rows):
        out_row = ppb_start + offset
        source_row = int(row["source_row"])
        for col in range(1, output_ws.max_column + 1):
            copy_cell(source_ws.cell(source_row, col), output_ws.cell(out_row, col))
        data_bar_rows.append(out_row)

    for offset, row in enumerate(rows):
        out_row = ppm_start + offset
        source_ppb_row = ppb_start + offset
        output_ws.cell(out_row, 1, row["dilution"])
        output_ws.cell(out_row, 2, sample if offset == 0 else None)
        output_ws.cell(out_row, 3, row["sample_name"])
        for col in range(1, ELEMENT_START_COL):
            copy_cell_style(source_ws.cell(int(row["source_row"]), col), output_ws.cell(out_row, col))

        for element_offset, _ in enumerate(elements):
            col = ELEMENT_START_COL + element_offset
            col_letter = get_column_letter(col)
            output_ws.cell(out_row, col, f"={col_letter}{source_ppb_row}/1000*$A{out_row}")
            copy_cell_style(source_ws.cell(int(row["source_row"]), col), output_ws.cell(out_row, col))
            output_ws.cell(out_row, col).fill = PatternFill(fill_type=None)
        data_bar_rows.append(out_row)

    output_ws.cell(selected_row, 1, "Selected concentration (ppm)")
    output_ws.cell(selected_row, 2, sample)
    for col in range(1, ELEMENT_START_COL):
        copy_cell_style(source_ws.cell(HEADER_ROW, col), output_ws.cell(selected_row, col))
        output_ws.cell(selected_row, col).fill = FILL_GOLD

    ppb_by_element = list(zip(*[row["ppb_values"] for row in rows]))
    dilution_values = [row["dilution"] for row in rows]
    for element_offset, ppb_values in enumerate(ppb_by_element):
        col = ELEMENT_START_COL + element_offset
        selected_idx = selected_index(list(ppb_values), dilution_values, calibration_range)
        selected_cell = output_ws.cell(selected_row, col)
        copy_cell_style(source_ws.cell(HEADER_ROW, col), selected_cell)
        selected_cell.fill = FILL_GOLD

        if selected_idx is None:
            selected_cell.value = None
            continue

        ppm_row = ppm_start + selected_idx
        col_letter = get_column_letter(col)
        selected_cell.value = f"={col_letter}{ppm_row}"

    apply_outline_border(output_ws, header_row, header_row, 1, last_col)
    apply_outline_border(output_ws, ppb_start, ppb_start + len(rows) - 1, 1, last_col)
    apply_outline_border(output_ws, ppm_start, ppm_start + len(rows) - 1, 1, last_col)
    apply_outline_border(output_ws, selected_row, selected_row, 1, last_col)
    section = (ppb_start, ppb_start + len(rows) - 1, ppm_start, ppm_start + len(rows) - 1, selected_row)
    return selected_row + 2, data_bar_rows, section


def apply_excel_formatting(
    workbook_path: Path,
    sheet_name: str,
    rows: list[int],
    selection_sections: list[tuple[int, int, int, int, int]],
    first_col: int,
    last_col: int,
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
) -> bool:
    workbook_literal = str(workbook_path.resolve()).replace("'", "''")
    sheet_literal = sheet_name.replace("'", "''")
    first_col_letter = get_column_letter(first_col)
    last_col_letter = get_column_letter(last_col)
    row_list = ",".join(str(row) for row in rows)
    section_list = ",".join(f"'{ppb_start},{ppb_end},{ppm_start},{ppm_end},{selected_row}'" for ppb_start, ppb_end, ppm_start, ppm_end, selected_row in selection_sections)
    orange_min = format_number(calibration_range.orange_min)
    green_min = format_number(calibration_range.green_min)
    green_max = format_number(calibration_range.green_max)

    script = f"""
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open('{workbook_literal}')
    $sheet = $workbook.Worksheets.Item('{sheet_literal}')
    $rows = @({row_list})
    foreach ($row in $rows) {{
        $rangeAddress = '{first_col_letter}' + $row + ':{last_col_letter}' + $row
        $range = $sheet.Range($rangeAddress)
        $bar = $range.FormatConditions.AddDatabar()
        $bar.AxisPosition = 0
        $bar.PercentMin = 0
        $bar.PercentMax = 100
        $bar.BarColor.Color = 13998939
        $bar.NegativeBarFormat.ColorType = 0
        $bar.NegativeBarFormat.Color.Color = 255
        $bar.ShowValue = $true
    }}
    $sections = @({section_list})
    foreach ($section in $sections) {{
        $parts = $section.Split(',')
        $ppbStart = [int]$parts[0]
        $ppbEnd = [int]$parts[1]
        $ppmStart = [int]$parts[2]
        $ppmEnd = [int]$parts[3]
        $selectedRow = [int]$parts[4]
        $ppbSectionRange = $sheet.Range('{first_col_letter}' + $ppbStart + ':{last_col_letter}' + $ppbEnd)
        $ppbHighlightRule = $ppbSectionRange.FormatConditions.Add(1, 1, '={green_min}', '={green_max}')
        $ppbHighlightRule.Interior.Color = {EXCEL_ICP_RANGE_FILL}
        for ($col = {first_col}; $col -le {last_col}; $col++) {{
            $colLetter = [string]$sheet.Cells.Item(1, $col).Address($false, $false)
            $colLetter = $colLetter -replace '1$', ''
            $ppbRange = $colLetter + '$' + $ppbStart + ':' + $colLetter + '$' + $ppbEnd
            $dilutionRange = '$A$' + $ppbStart + ':$A$' + $ppbEnd
            $topPpbCell = $colLetter + $ppbStart
            $topDilutionCell = '$A' + $ppbStart
            $ppmRange = $sheet.Range($colLetter + $ppmStart + ':' + $colLetter + $ppmEnd)
            $selectedCell = $sheet.Range($colLetter + $selectedRow)

            $formulaGreen = '=AND(' + $topPpbCell + '>={green_min},' + $topPpbCell + '<={green_max},' + $topDilutionCell + '=MINIFS(' + $dilutionRange + ',' + $ppbRange + ',">={green_min}",' + $ppbRange + ',"<={green_max}"))'
            $rule = $ppmRange.FormatConditions.Add(2, [Type]::Missing, $formulaGreen)
            $rule.Interior.Color = {EXCEL_FILL_GREEN}

            $formulaOrange = '=AND(COUNTIFS(' + $ppbRange + ',">={green_min}",' + $ppbRange + ',"<={green_max}")=0,' + $topPpbCell + '>={orange_min},' + $topPpbCell + '<{green_min},' + $topDilutionCell + '=MINIFS(' + $dilutionRange + ',' + $ppbRange + ',">={orange_min}",' + $ppbRange + ',"<{green_min}"))'
            $rule = $ppmRange.FormatConditions.Add(2, [Type]::Missing, $formulaOrange)
            $rule.Interior.Color = {EXCEL_FILL_ORANGE}

            $formulaGreyLow = '=AND(MAX(' + $ppbRange + ')<{orange_min},' + $topDilutionCell + '=MIN(' + $dilutionRange + '))'
            $rule = $ppmRange.FormatConditions.Add(2, [Type]::Missing, $formulaGreyLow)
            $rule.Interior.Color = {EXCEL_FILL_GREY}

            $formulaRedHigh = '=AND(MIN(' + $ppbRange + ')>{green_max},' + $topDilutionCell + '=MAX(' + $dilutionRange + '))'
            $rule = $ppmRange.FormatConditions.Add(2, [Type]::Missing, $formulaRedHigh)
            $rule.Interior.Color = {EXCEL_FILL_RED}

            $formulaSelectedGreen = '=COUNTIFS(' + $ppbRange + ',">={green_min}",' + $ppbRange + ',"<={green_max}")>0'
            $rule = $selectedCell.FormatConditions.Add(2, [Type]::Missing, $formulaSelectedGreen)
            $rule.Interior.Color = {EXCEL_FILL_GREEN}

            $formulaSelectedOrange = '=AND(COUNTIFS(' + $ppbRange + ',">={green_min}",' + $ppbRange + ',"<={green_max}")=0,COUNTIFS(' + $ppbRange + ',">={orange_min}",' + $ppbRange + ',"<{green_min}")>0)'
            $rule = $selectedCell.FormatConditions.Add(2, [Type]::Missing, $formulaSelectedOrange)
            $rule.Interior.Color = {EXCEL_FILL_ORANGE}

            $formulaSelectedGrey = '=MAX(' + $ppbRange + ')<{orange_min}'
            $rule = $selectedCell.FormatConditions.Add(2, [Type]::Missing, $formulaSelectedGrey)
            $rule.Interior.Color = {EXCEL_FILL_GREY}

            $formulaSelectedRed = '=MIN(' + $ppbRange + ')>{green_max}'
            $rule = $selectedCell.FormatConditions.Add(2, [Type]::Missing, $formulaSelectedRed)
            $rule.Interior.Color = {EXCEL_FILL_RED}
        }}
    }}
    $workbook.Save()
    exit 0
}} catch {{
    Write-Error $_.Exception.Message
    exit 1
}} finally {{
    if ($workbook -ne $null) {{ $workbook.Close($true) }}
    if ($excel -ne $null) {{ $excel.Quit() }}
}}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print("WARNING: Excel data bars could not be applied.")
        if result.stderr.strip():
            print(result.stderr.strip())
        return False
    return True


def resolve_source_sheet(source_wb, requested_sheet: str | None) -> str:
    """Find the ICP sheet to read, optionally honoring an exact user request."""

    if requested_sheet:
        if requested_sheet not in source_wb.sheetnames:
            available = ", ".join(source_wb.sheetnames)
            raise ValueError(f'Sheet "{requested_sheet}" was not found. Available sheets: {available}')
        return requested_sheet

    icp_sheets = [name for name in source_wb.sheetnames if name.upper().endswith(" ICP")]
    if len(icp_sheets) == 1:
        return icp_sheets[0]
    if DEFAULT_SOURCE_SHEET_NAME in source_wb.sheetnames:
        return DEFAULT_SOURCE_SHEET_NAME
    if icp_sheets:
        available = ", ".join(icp_sheets)
        raise ValueError(f"Multiple ICP sheets found ({available}). Use --sheet to choose one.")

    available = ", ".join(source_wb.sheetnames)
    raise ValueError(f'No sheet ending with " ICP" was found. Available sheets: {available}')


def apply_openpyxl_formatting_fallback(
    workbook_path: Path,
    sheet_name: str,
    rows: list[int],
    selection_sections,
    first_col: int,
    last_col: int,
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
) -> bool:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    
    ws.conditional_formatting._cf_rules.clear()

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    orange_fill = PatternFill("solid", fgColor="F4B183")
    grey_fill = PatternFill("solid", fgColor="D9D9D9")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    selected_fill = PatternFill("solid", fgColor="FFD966")
    
    data_bar_rule = DataBarRule(
        start_type="min",
        end_type="max",
        color="5B9BD5",
        showValue=True,
    )
    
    for (ppb_start, ppb_end, _, _, _) in selection_sections:

        for row in range(ppb_start, ppb_end + 1):
            for col in range(first_col, last_col + 1):
                cell = ws.cell(row=row, column=col)
                value = cell.value

                if (
                    isinstance(value, (int, float))
                    and calibration_range.green_min <= value <= calibration_range.green_max
                ):
                    cell.fill = green_fill
                    
            start = f"{get_column_letter(first_col)}{row}"
            end = f"{get_column_letter(last_col)}{row}"
            cell_range = f"{start}:{end}"

            ws.conditional_formatting.add(cell_range, data_bar_rule)

    # Highlight "Selected concentration"
    for (ppb_start, ppb_end, ppm_start, ppm_end, selected_row) in selection_sections:
        
        selected_ppm_row_by_col = {}

        for col in range(first_col, last_col + 1):

            ppb_values = []
            for r in range(ppb_start, ppb_end + 1):
                val = ws.cell(row=r, column=col).value
                dilution = parse_number(ws.cell(row=r, column=1).value)
                if isinstance(val, (int, float)) and dilution is not None:
                    ppb_values.append((r, val, dilution))

            if not ppb_values:
                continue

            green_values = [
                (r, v, dilution)
                for r, v, dilution in ppb_values
                if calibration_range.green_min <= v <= calibration_range.green_max
            ]
            orange_values = [
                (r, v, dilution)
                for r, v, dilution in ppb_values
                if calibration_range.orange_min <= v < calibration_range.green_min
            ]

            if green_values:
                chosen_row, _, _ = min(green_values, key=lambda x: x[2])
                fill = green_fill
            elif orange_values:
                chosen_row, _, _ = min(orange_values, key=lambda x: x[2])
                fill = orange_fill
            elif all(v < calibration_range.orange_min for _, v, _ in ppb_values):
                chosen_row, _, _ = min(ppb_values, key=lambda x: x[2])
                fill = grey_fill
            elif all(v > calibration_range.green_max for _, v, _ in ppb_values):
                chosen_row, _, _ = max(ppb_values, key=lambda x: x[2])
                fill = red_fill
            else:
                continue

            ppm_row = ppm_start + (chosen_row - ppb_start)
            ws.cell(row=ppm_row, column=col).fill = fill
            selected_ppm_row_by_col[col] = ppm_row
    
        # gold only for labels
        ws.cell(row=selected_row, column=1).fill = selected_fill
        ws.cell(row=selected_row, column=2).fill = selected_fill
        ws.cell(row=selected_row, column=3).fill = selected_fill
        
        # copy per-element color from PPM row
        for col, ppm_row in selected_ppm_row_by_col.items():
            ppm_cell = ws.cell(row=ppm_row, column=col)
            selected_cell = ws.cell(row=selected_row, column=col)

            if ppm_cell.fill and ppm_cell.fill.fill_type == "solid":
                selected_cell.fill = copy(ppm_cell.fill)

    wb.save(workbook_path)
    return True


def build_concentration_workbook(
    source_path: Path,
    output_path: Path,
    source_sheet_name: str | None = None,
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
) -> None:
    source_wb = load_workbook(source_path, data_only=False)
    resolved_source_sheet = resolve_source_sheet(source_wb, source_sheet_name)
    source_ws = source_wb[resolved_source_sheet]
    elements, groups = read_groups(source_ws)

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = OUTPUT_SHEET_NAME
    output_ws.freeze_panes = "D3"

    for col_idx in range(1, source_ws.max_column + 1):
        letter = get_column_letter(col_idx)
        output_ws.column_dimensions[letter].width = source_ws.column_dimensions[letter].width or 14

    current_row = 1
    data_bar_rows: list[int] = []
    selection_sections: list[tuple[int, int, int, int, int]] = []
    for sample, rows in groups.items():
        current_row, block_data_bar_rows, section = write_sample_block(
            output_ws,
            source_ws,
            current_row,
            sample,
            rows,
            elements,
            calibration_range,
        )
        data_bar_rows.extend(block_data_bar_rows)
        selection_sections.append(section)

    output_wb.save(output_path)
    if excel_backend() == "windows":
        excel_formatting_applied = apply_excel_formatting(
            output_path,
            OUTPUT_SHEET_NAME,
            data_bar_rows,
            selection_sections,
            ELEMENT_START_COL,
            ELEMENT_START_COL + len(elements) - 1,
            calibration_range,
        )
    else:
        excel_formatting_applied = apply_openpyxl_formatting_fallback(
            output_path,
            OUTPUT_SHEET_NAME,
            data_bar_rows,
            selection_sections,
            ELEMENT_START_COL,
            ELEMENT_START_COL + len(elements) - 1,
            calibration_range,
        )

    print(f"Concentration workbook saved: {output_path}")
    print(f"Source ICP sheet used: {resolved_source_sheet}")
    print(f"Sample groups written: {len(groups)}")
    print(f"Elements written: {len(elements)}")
    print(f"Excel conditional formatting applied: {'yes' if excel_formatting_applied else 'no'}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate original-solution ICP concentration workbook from an ICP sheet.")
    parser.add_argument("--source", type=Path, required=True, help='Workbook containing an "<initials> ICP" sheet')
    parser.add_argument("--output", type=Path, required=True, help="Output concentration workbook")
    parser.add_argument(
        "--sheet",
        help='Source ICP sheet name, such as "SH ICP", "T ICP", or "H ICP". If omitted, the only "* ICP" sheet is used.',
    )
    parser.add_argument("--orange-min", type=float, default=DEFAULT_CALIBRATION_RANGE.orange_min, help="Minimum ppb for the orange lower-confidence calibration range.")
    parser.add_argument("--green-min", type=float, default=DEFAULT_CALIBRATION_RANGE.green_min, help="Minimum ppb for the green accurate calibration range.")
    parser.add_argument("--green-max", type=float, default=DEFAULT_CALIBRATION_RANGE.green_max, help="Maximum ppb for the green accurate calibration range.")
    args = parser.parse_args()

    calibration_range = CalibrationRange(
        orange_min=args.orange_min,
        green_min=args.green_min,
        green_max=args.green_max,
    )
    build_concentration_workbook(args.source, args.output, args.sheet, calibration_range)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
