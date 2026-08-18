"""Process raw ICP exports into cleaned and highlighted workbooks.

The script validates the expected ICP template before doing any work. It then
extracts only analyte concentration columns with pandas and uses openpyxl to
create these sheets:
    - original raw data
    - cleaned highlighted all-sample data
    - <initials> samples
    - <initials> ICP, with editable dilution/sample columns and row-wise data bars

Usage:
    python process_icp_data.py input.xlsx --output cleaned.xlsx
    python process_icp_data.py input.csv --output cleaned.xlsx
    python process_icp_data.py input.csv --initials T --output T_cleaned.xlsx
"""

from __future__ import annotations

__author__ = "Shihua Han"
__version__ = "0.4.1"

import argparse
import re
import subprocess
from copy import copy
from pathlib import Path
from typing import Iterable, NamedTuple

import pandas as pd
from calibration_settings import DEFAULT_CALIBRATION_RANGE, DEFAULT_HIGHLIGHT_THRESHOLDS, CalibrationRange, HighlightThresholds, format_number
from excel_backend import excel_backend
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Fixed template positions. Change these only if the lab template changes.
HEADER_ROW = 1
SUBHEADER_ROW = 2
RAW_DATA_START_ROW = 3
DEFAULT_SAMPLE_INITIALS = "SH"
PROCESSED_TABLE_START_COL = 4  # Column D
PROCESSED_HEADER_ROW = 1
PROCESSED_SUBHEADER_ROW = 2
PROCESSED_DATA_START_ROW = 3

MEASURED_CONC_LABEL_PATTERN = re.compile(r"^(?:meas\.|measured)\s*conc\.(?:\s*\[.*\])?$", re.IGNORECASE)

EXCEL_ICP_RANGE_FILL = 13693658

FILL_ORANGE = PatternFill("solid", fgColor="F4B183")
FILL_DARK_ORANGE = PatternFill("solid", fgColor="C65911")


class ElementBlock(NamedTuple):
    """Raw worksheet element block discovered from the two-row ICP header."""

    name: str
    start_col_idx: int  # zero-based pandas column index
    conc_col_idx: int  # zero-based pandas column index


class ValidationResult(NamedTuple):
    ok: bool
    errors: list[str]
    element_failures: list[str]
    sample_name_col_idx: int | None
    element_blocks: list[ElementBlock]
    sh_row_indices: list[int]  # zero-based pandas row indices


def normalize_sample_initials(value: str | None) -> str:
    """Return a clean sample prefix such as SH, T, or H."""

    initials = (value or DEFAULT_SAMPLE_INITIALS).strip()
    if not initials:
        raise ValueError("Sample initials cannot be blank.")
    if not re.fullmatch(r"[A-Za-z]+", initials):
        raise ValueError("Sample initials must contain letters only, such as SH, T, or H.")
    return initials.upper()


def sample_name_pattern(sample_initials: str) -> re.Pattern[str]:
    return re.compile(rf"^{re.escape(sample_initials)}_?[A-Za-z0-9]+$", re.IGNORECASE)


def sample_sheet_name(sample_initials: str) -> str:
    return f"{sample_initials} samples"


def icp_sheet_name(sample_initials: str) -> str:
    return f"{sample_initials} ICP"


def normalize_header(value: object) -> str:
    """Return a trimmed string for matching headers while tolerating blanks."""

    if pd.isna(value):
        return ""
    return str(value).strip()


def parse_number(value: object) -> float | None:
    """Convert numeric cells and text-formatted numbers to float.

    Handles blanks, commas, Excel text, and values such as '<0.000' by returning
    the numeric part with the sign implied by the text.
    """

    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text.upper() in {"N/A", "NA", "NULL", "NONE"}:
        return None

    text = text.replace(",", "")
    less_than = text.startswith("<")
    if less_than:
        text = text[1:].strip()

    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text)
    if not match:
        return None

    number = float(match.group(0))
    return min(number, 0.0) if less_than else number


def parse_nominal_ppb(sample_name: object) -> float | None:
    """Return the nominal ppb value embedded in a calibration/sample name."""

    text = normalize_header(sample_name)
    match = re.search(r"([-+]?\d*\.?\d+)\s*ppb", text, re.IGNORECASE)
    return float(match.group(1)) if match else None


def is_blank_standard(sample_name: object) -> bool:
    """Return True for blank/blk internal-standard sample names."""

    text = normalize_header(sample_name).lower()
    return bool(re.search(r"\b(?:blk|blank)\b", text))


def safe_sheet_name(name: str) -> str:
    """Return an Excel-safe worksheet name capped at Excel's 31-char limit."""

    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip()
    return (cleaned or "Sheet")[:31]


def find_last_used_col(df: pd.DataFrame) -> int:
    """Return the zero-based index of the last column with any nonblank value."""

    for col_idx in range(df.shape[1] - 1, -1, -1):
        if df.iloc[:, col_idx].map(normalize_header).ne("").any():
            return col_idx
    return -1


def find_sample_name_column(subheaders: list[str]) -> int | None:
    """Locate the Sample Name column from row 2 subheaders."""

    for idx, label in enumerate(subheaders):
        if label.lower() == "sample name":
            return idx
    return None


def element_symbol(element_name: str) -> str:
    """Return the element symbol from a raw ICP header.

    Examples:
        "7  Li  [ No Gas ]" -> "Li"
        "56  Fe  [ H2 ]" -> "Fe"
    """

    match = re.match(r"^\s*\d+\s+([A-Z][a-z]?)\b", element_name)
    if match:
        return match.group(1)

    # Fallback for headers that already start with a symbol or have no mass.
    match = re.search(r"\b([A-Z][a-z]?)\b", element_name)
    return match.group(1) if match else element_name.strip()


def discover_element_blocks(headers: list[str], subheaders: list[str], last_col: int) -> tuple[list[ElementBlock], list[str]]:
    """Validate and return element measured-concentration columns.

    Raw ICP exports can vary in group width. Instead of assuming every element
    occupies exactly three columns, locate each ``Meas. Conc.`` subheader and
    associate it with the nearest element name to its left in row 1.
    """

    errors: list[str] = []
    blocks: list[ElementBlock] = []

    measured_conc_cols = [
        idx
        for idx, label in enumerate(subheaders[: last_col + 1])
        if MEASURED_CONC_LABEL_PATTERN.match(label)
    ]
    if not measured_conc_cols:
        errors.append('No measured concentration columns found in row 2; expected "Meas. Conc." or "Measured Conc.".')
        return blocks, errors

    for conc_col_idx in measured_conc_cols:
        element_start_col_idx = next(
            (idx for idx in range(conc_col_idx, -1, -1) if headers[idx]),
            None,
        )
        excel_col = get_column_letter(conc_col_idx + 1)

        if element_start_col_idx is None:
            errors.append(f"Measured concentration column {excel_col} has no element name to its left in row 1.")
            continue

        element_name = headers[element_start_col_idx]
        group = subheaders[element_start_col_idx : conc_col_idx + 1]
        if "ISTD" in element_name.upper() or any("ISTD RECOVERY" in label.upper() for label in group):
            # Internal-standard groups are instrument QC fields, not analyte
            # concentration columns for cleaned analyte tables.
            continue

        blocks.append(ElementBlock(element_name, element_start_col_idx, conc_col_idx))

    if not blocks:
        errors.append("No analyte measured concentration columns were found after excluding internal-standard fields.")

    return blocks, errors


def validate_raw_layout(df: pd.DataFrame, sample_initials: str) -> ValidationResult:
    """Validate the ICP worksheet before changing or saving anything."""

    errors: list[str] = []
    element_failures: list[str] = []

    if df.shape[0] < 2:
        return ValidationResult(False, ["Worksheet must contain at least two header rows."], [], None, [], [])

    last_col = find_last_used_col(df)
    if last_col < 0:
        return ValidationResult(False, ["Worksheet appears to be empty."], [], None, [], [])

    headers = [normalize_header(value) for value in df.iloc[HEADER_ROW - 1, : last_col + 1].tolist()]
    subheaders = [normalize_header(value) for value in df.iloc[SUBHEADER_ROW - 1, : last_col + 1].tolist()]

    if not any(headers):
        errors.append("Row 1 does not contain element names.")
    if not any(MEASURED_CONC_LABEL_PATTERN.match(label) for label in subheaders):
        errors.append('Row 2 does not contain repeated subheaders including "Meas. Conc." or "Measured Conc.".')
    if PROCESSED_TABLE_START_COL != 4:
        errors.append("Processed table start column is not configured as column D.")

    sample_name_col_idx = find_sample_name_column(subheaders)
    if sample_name_col_idx is None:
        errors.append('Sample name column was not found in row 2; expected subheader "Sample Name".')
        sh_row_indices: list[int] = []
    else:
        sample_values = df.iloc[RAW_DATA_START_ROW - 1 :, sample_name_col_idx].map(normalize_header)
        sh_mask = sample_values.str.match(sample_name_pattern(sample_initials), na=False)
        sh_row_indices = sample_values[sh_mask].index.tolist()
        if not sh_row_indices:
            errors.append(
                f'No {sample_initials} sample rows found. Expected names such as '
                f'"{sample_initials}1", "{sample_initials}_10k", or "{sample_initials}_x10k".'
            )

    element_blocks, block_errors = discover_element_blocks(headers, subheaders, last_col)
    errors.extend(block_errors)

    if sh_row_indices and element_blocks:
        for block in element_blocks:
            values = [parse_number(df.iat[row_idx, block.conc_col_idx]) for row_idx in sh_row_indices]
            if all(value is None for value in values):
                element_failures.append(block.name)

    return ValidationResult(
        ok=not errors,
        errors=errors,
        element_failures=element_failures,
        sample_name_col_idx=sample_name_col_idx,
        element_blocks=element_blocks,
        sh_row_indices=sh_row_indices,
    )


def load_input_dataframe(path: Path, sheet_name: str | int | None) -> pd.DataFrame:
    """Read the raw worksheet with pandas without treating any row as a header."""

    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, header=None, dtype=object, keep_default_na=False)
    return pd.read_excel(path, sheet_name=sheet_name or 0, header=None, dtype=object, engine="openpyxl")


def load_or_create_workbook(
    path: Path,
    sheet_name: str | int | None,
    source_name: str | None = None,
) -> tuple[Workbook, str, bool]:
    """Open an Excel workbook or create one from CSV data.

    Returns the workbook, source sheet name, and whether the source was CSV.
    """

    if path.suffix.lower() == ".csv":
        wb = Workbook()
        ws = wb.active
        displayed_source_name = Path(source_name).stem if source_name else path.stem
        ws.title = safe_sheet_name(displayed_source_name)
        csv_df = pd.read_csv(path, header=None, dtype=object, keep_default_na=False)
        for row in csv_df.itertuples(index=False, name=None):
            ws.append(list(row))
        return wb, ws.title, True

    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, keep_vba=keep_vba)
    if isinstance(sheet_name, str):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'Sheet "{sheet_name}" not found. Available sheets: {", ".join(wb.sheetnames)}')
        source_sheet_name = sheet_name
    elif isinstance(sheet_name, int):
        source_sheet_name = wb.sheetnames[sheet_name]
    else:
        source_sheet_name = wb.sheetnames[0]
    return wb, source_sheet_name, False


def copy_cell_style(source_cell, target_cell) -> None:
    """Copy basic cell formatting while leaving values/formulas independent."""

    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def prepare_sheet(wb: Workbook, sheet_name: str) -> object:
    """Create a fresh sheet, replacing a prior generated sheet of the same name."""

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    return wb.create_sheet(sheet_name)


def fill_for_percent_error(
    value: float | None,
    nominal: float | None,
    highlight_thresholds: HighlightThresholds = DEFAULT_HIGHLIGHT_THRESHOLDS,
) -> PatternFill | None:
    """Return the orange fill for errors beyond relative-error thresholds."""

    if value is None or nominal is None:
        return None
    percent_error = abs(value - nominal) / abs(nominal) * 100
    if percent_error > highlight_thresholds.internal_orange:
        return FILL_DARK_ORANGE
    if percent_error > highlight_thresholds.internal_light_orange:
        return FILL_ORANGE
    return None


def fill_for_blank_standard(
    value: float | None,
    highlight_thresholds: HighlightThresholds = DEFAULT_HIGHLIGHT_THRESHOLDS,
) -> PatternFill | None:
    """Return the orange fill for blank values beyond absolute thresholds."""

    if value is None:
        return None

    absolute_value = abs(value)
    if absolute_value > highlight_thresholds.blank_orange:
        return FILL_DARK_ORANGE
    if absolute_value > highlight_thresholds.blank_light_orange:
        return FILL_ORANGE
    return None


def highlight_internal_standard_rows(
    cleaned_ws,
    sample_output_col: int,
    element_start_col: int,
    highlight_thresholds: HighlightThresholds = DEFAULT_HIGHLIGHT_THRESHOLDS,
) -> int:
    """Highlight ppb/blank standard rows in a cleaned concentration worksheet.

    For rows whose Sample Name contains a nominal ppb value, analyte concentration
    cells are colored by relative-error thresholds. Blank/blk rows use absolute
    concentration thresholds.
    """

    highlighted_count = 0

    for row_idx in range(PROCESSED_DATA_START_ROW, cleaned_ws.max_row + 1):
        sample_name = cleaned_ws.cell(row_idx, sample_output_col).value
        nominal_ppb = parse_nominal_ppb(sample_name)
        blank_row = nominal_ppb is None and is_blank_standard(sample_name)
        if nominal_ppb is None and not blank_row:
            continue

        for col_idx in range(element_start_col, cleaned_ws.max_column + 1):
            cell = cleaned_ws.cell(row_idx, col_idx)
            value = parse_number(cell.value)
            fill = (
                fill_for_blank_standard(value, highlight_thresholds)
                if blank_row
                else fill_for_percent_error(value, nominal_ppb, highlight_thresholds)
            )
            if fill is not None:
                cell.fill = fill
                highlighted_count += 1

    return highlighted_count


def write_cleaned_table(
    wb: Workbook,
    raw_sheet_name: str,
    df: pd.DataFrame,
    validation: ValidationResult,
    sheet_name: str,
    row_indices: Iterable[int],
) -> object:
    """Create a cleaned concentration sheet from selected raw row indices."""

    raw_ws = wb[raw_sheet_name]
    processed_ws = prepare_sheet(wb, sheet_name)

    sample_col = validation.sample_name_col_idx
    if sample_col is None:
        raise ValueError("Cannot write processed sheet without a validated Sample Name column.")

    # Keep only Sample Name from the metadata area. It lives in raw column G and
    # is written to column C so analyte columns still start at processed column D.
    sample_output_col = PROCESSED_TABLE_START_COL - 1
    processed_ws.cell(PROCESSED_HEADER_ROW, sample_output_col, "Sample Name")
    processed_ws.cell(PROCESSED_SUBHEADER_ROW, sample_output_col, "Sample Name")

    for element_offset, block in enumerate(validation.element_blocks):
        out_col = PROCESSED_TABLE_START_COL + element_offset
        processed_ws.cell(PROCESSED_HEADER_ROW, out_col, element_symbol(block.name))
        processed_ws.cell(PROCESSED_SUBHEADER_ROW, out_col, "Measured Conc. [ ppb ]")

    data = df.loc[list(row_indices)].copy()
    for out_row_offset, (source_row_idx, row) in enumerate(data.iterrows(), start=PROCESSED_DATA_START_ROW):
        processed_ws.cell(out_row_offset, sample_output_col, row.iloc[sample_col])
        if sample_col + 1 <= raw_ws.max_column:
            copy_cell_style(
                raw_ws.cell(source_row_idx + 1, sample_col + 1),
                processed_ws.cell(out_row_offset, sample_output_col),
            )

        for element_offset, block in enumerate(validation.element_blocks):
            out_col = PROCESSED_TABLE_START_COL + element_offset
            value = row.iloc[block.conc_col_idx]
            processed_ws.cell(out_row_offset, out_col, value)

    for col_idx in range(1, processed_ws.max_column + 1):
        processed_ws.column_dimensions[get_column_letter(col_idx)].width = 14
    processed_ws.freeze_panes = "D3"

    return processed_ws


def create_icp_sheet_from_samples(wb: Workbook, samples_ws, sheet_name: str):
    """Copy the sample sheet, normalize numbers, and add row data bars."""

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    icp_ws = wb.copy_worksheet(samples_ws)
    icp_ws.title = sheet_name

    icp_ws.cell(PROCESSED_HEADER_ROW, 1, "Dilution factor")
    icp_ws.cell(PROCESSED_SUBHEADER_ROW, 1, "Dilution factor")
    icp_ws.cell(PROCESSED_HEADER_ROW, 2, "Sample")
    icp_ws.cell(PROCESSED_SUBHEADER_ROW, 2, "Sample")
    icp_ws.cell(PROCESSED_HEADER_ROW, 1).comment = Comment(
        "Enter the dilution factor for each sample row below.", "Codex"
    )
    icp_ws.cell(PROCESSED_HEADER_ROW, 2).comment = Comment(
        "Enter the sample identifier/name for each sample row below.", "Codex"
    )
    icp_ws.column_dimensions["A"].width = 16
    icp_ws.column_dimensions["B"].width = 18
    for row_idx in range(PROCESSED_DATA_START_ROW, icp_ws.max_row + 1):
        icp_ws.cell(row_idx, 1, None)
        icp_ws.cell(row_idx, 2, None)

    element_start_col = PROCESSED_TABLE_START_COL
    element_end_col = icp_ws.max_column

    for row_idx in range(PROCESSED_DATA_START_ROW, icp_ws.max_row + 1):
        for col_idx in range(element_start_col, element_end_col + 1):
            cell = icp_ws.cell(row_idx, col_idx)
            number = parse_number(cell.value)
            if number is None:
                continue

            # Store numeric values in this derived sheet so Excel data bars work.
            cell.value = number

        row_range = (
            f"{get_column_letter(element_start_col)}{row_idx}:"
            f"{get_column_letter(element_end_col)}{row_idx}"
        )
        icp_ws.conditional_formatting.add(
            row_range,
            DataBarRule(start_type="min", end_type="max", color="5B9BD5", showValue=True),
        )

    return icp_ws


def apply_excel_automatic_data_bars(
    workbook_path: Path,
    sheet_name: str,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
) -> bool:
    """Use Excel COM to set row-wise data bars to Excel's Automatic min/max.

    Openpyxl can create data bars, but its public API exposes min/max threshold
    types rather than Excel's UI setting named Automatic. This optional post-save
    pass uses local Excel when available so the generated rule matches the manual
    Excel workflow: New Rule -> Data Bar -> Minimum Automatic, Maximum Automatic.
    """

    workbook_literal = str(workbook_path.resolve()).replace("'", "''")
    sheet_literal = sheet_name.replace("'", "''")
    first_col_letter = get_column_letter(first_col)
    last_col_letter = get_column_letter(last_col)
    green_min = format_number(calibration_range.green_min)
    green_max = format_number(calibration_range.green_max)
    # Excel constants: xlConditionValueAutomatic = 7.
    script = f"""
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open('{workbook_literal}')
    $sheet = $workbook.Worksheets.Item('{sheet_literal}')
    for ($row = {first_row}; $row -le {last_row}; $row++) {{
        $rangeAddress = '{first_col_letter}' + $row + ':{last_col_letter}' + $row
        $range = $sheet.Range($rangeAddress)
        $range.FormatConditions.Delete()
        $bar = $range.FormatConditions.AddDatabar()
        # AddDatabar creates Excel's default Automatic min/max points. Leave
        # those points untouched so the UI shows Automatic for both ends.
        $bar.AxisPosition = 0
        $bar.PercentMin = 0
        $bar.PercentMax = 100
        $bar.BarColor.Color = 13998939
        $bar.NegativeBarFormat.ColorType = 0
        $bar.NegativeBarFormat.Color.Color = 255
        $bar.ShowValue = $true
    }}
    $highlightRange = $sheet.Range('{first_col_letter}{first_row}:{last_col_letter}{last_row}')
    $rule = $highlightRange.FormatConditions.Add(1, 1, '={green_min}', '={green_max}')
    $rule.Interior.Color = {EXCEL_ICP_RANGE_FILL}
    $workbook.Save()
    exit 0
}} catch {{
    Write-Error $_.Exception.Message
    exit 1
}} finally {{
    if ($workbook -ne $null) {{ $workbook.Close($true) }}
    if ($excel -ne $null) {{ $excel.Quit() }}
}}
"""

    result = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print("WARNING: Excel Automatic data bars could not be applied; kept openpyxl data bars.")
        if result.stderr.strip():
            print(result.stderr.strip())
        return False
    return True
    

def apply_openpyxl_data_bars_fallback(
    workbook_path: Path,
    sheet_name: str,
    first_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
) -> bool:
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]
    
    ws.conditional_formatting._cf_rules.clear()

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )
        
    # Styles
    data_bar_rule = DataBarRule(
        start_type="min",
        end_type="max",
        color="5B9BD5",
        showValue=True,
    )

    # Apply row-wise
    for row in range(first_row, last_row + 1):
        start = f"{get_column_letter(first_col)}{row}"
        end = f"{get_column_letter(last_col)}{row}"
        cell_range = f"{start}:{end}"

        # Data bars
        ws.conditional_formatting.add(cell_range, data_bar_rule)

        # Accurate calibration-range highlight
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="between",
                formula=[
                    format_number(calibration_range.green_min),
                    format_number(calibration_range.green_max),
                ],
                fill=green_fill
            )
        )

    wb.save(workbook_path)
    return True
    

def default_output_path(input_path: Path, sample_initials: str) -> Path:
    """Return a non-destructive output filename next to the input file."""

    suffix = ".xlsm" if input_path.suffix.lower() == ".xlsm" else ".xlsx"
    return input_path.with_name(f"{input_path.stem}_highlighted_{sample_initials}_samples{suffix}")


def process_icp_file(
    input_path: Path,
    output_path: Path | None,
    sheet_name: str | int | None,
    sample_initials: str,
    calibration_range: CalibrationRange = DEFAULT_CALIBRATION_RANGE,
    highlight_thresholds: HighlightThresholds = DEFAULT_HIGHLIGHT_THRESHOLDS,
    source_name: str | None = None,
) -> int:
    """Validate, process, save, and print a concise run summary.

    ``source_name`` supplies a user-facing filename when ``input_path`` is a
    temporary upload file.
    """

    df = load_input_dataframe(input_path, sheet_name)
    validation = validate_raw_layout(df, sample_initials)

    if not validation.ok:
        print("ERROR: ICP worksheet layout validation failed. Processing stopped.")
        for error in validation.errors:
            print(f"- {error}")
        if validation.element_failures:
            print("Warning: missing concentration values were also found for these elements:")
            for element in validation.element_failures:
                print(f"- {element}")
        return 1

    wb, raw_sheet_name, _ = load_or_create_workbook(input_path, sheet_name, source_name)
    displayed_source_name = Path(source_name).stem if source_name else input_path.stem
    highlighted_suffix = " highlighted"
    # Excel sheet names are limited to 31 characters. Reserve room for the
    # suffix so long upload names still clearly identify this sheet's purpose.
    highlighted_sheet_name = safe_sheet_name(
        f"{displayed_source_name[:31 - len(highlighted_suffix)]}{highlighted_suffix}"
    )
    all_data_row_indices = range(RAW_DATA_START_ROW - 1, len(df))
    highlighted_ws = write_cleaned_table(wb, raw_sheet_name, df, validation, highlighted_sheet_name, all_data_row_indices)
    highlighted_count = highlight_internal_standard_rows(
        highlighted_ws,
        sample_output_col=PROCESSED_TABLE_START_COL - 1,
        element_start_col=PROCESSED_TABLE_START_COL,
        highlight_thresholds=highlight_thresholds,
    )
    processed_ws = write_cleaned_table(wb, raw_sheet_name, df, validation, sample_sheet_name(sample_initials), validation.sh_row_indices)
    icp_ws = create_icp_sheet_from_samples(wb, processed_ws, icp_sheet_name(sample_initials))

    save_path = output_path or default_output_path(input_path, sample_initials)
    wb.save(save_path)
    if excel_backend() == "windows":
        automatic_bars_applied = apply_excel_automatic_data_bars(
            save_path,
            sheet_name=icp_ws.title,
            first_row=PROCESSED_DATA_START_ROW,
            last_row=icp_ws.max_row,
            first_col=PROCESSED_TABLE_START_COL,
            last_col=icp_ws.max_column,
            calibration_range=calibration_range,
        )
    else:
        automatic_bars_applied = apply_openpyxl_data_bars_fallback(
            save_path,
            sheet_name=icp_ws.title,
            first_row=PROCESSED_DATA_START_ROW,
            last_row=icp_ws.max_row,
            first_col=PROCESSED_TABLE_START_COL,
            last_col=icp_ws.max_column,
            calibration_range=calibration_range,
        )

    print(f"Processed element count: {len(validation.element_blocks)}")
    print(f"{sample_initials} sample rows kept: {len(validation.sh_row_indices)}")
    print(f"ICP sheet created: {icp_ws.title}")
    print(f"Excel Automatic data bars applied: {'yes' if automatic_bars_applied else 'no'}")
    print(f"Internal-standard cells highlighted: {highlighted_count}")
    if validation.element_failures:
        print("Warning: concentration values are missing for these elements in the selected sample rows; blanks were left unchanged:")
        for element in sorted(set(validation.element_failures)):
            print(f"- {element}")
    else:
        print("Missing concentration warnings: none")
    print(f"Saved processed workbook: {save_path}")
    return 0


def parse_sheet_arg(value: str | None) -> str | int | None:
    """Allow --sheet to be either a sheet name or zero-based numeric index."""

    if value is None:
        return None
    if value.isdigit():
        return int(value)
    return value


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate and process fixed-template ICP data exports.")
    parser.add_argument("input", type=Path, help="Input .xlsx, .xlsm, or .csv ICP export")
    parser.add_argument("--output", "-o", type=Path, help="Output workbook path; defaults to *_processed.xlsx")
    parser.add_argument("--sheet", help="Worksheet name or zero-based sheet index for Excel inputs")
    parser.add_argument(
        "--initials",
        default=DEFAULT_SAMPLE_INITIALS,
        help='Sample-name prefix to keep, such as "SH", "T", "H", or "CS". Defaults to SH.',
    )
    parser.add_argument("--orange-min", type=float, default=DEFAULT_CALIBRATION_RANGE.orange_min, help="Minimum ppb for the orange lower-confidence calibration range.")
    parser.add_argument("--green-min", type=float, default=DEFAULT_CALIBRATION_RANGE.green_min, help="Minimum ppb for the green accurate calibration range.")
    parser.add_argument("--green-max", type=float, default=DEFAULT_CALIBRATION_RANGE.green_max, help="Maximum ppb for the green accurate calibration range.")
    parser.add_argument("--internal-light-orange", type=float, default=DEFAULT_HIGHLIGHT_THRESHOLDS.internal_light_orange, help="Internal-standard percent-error tolerance before orange highlighting.")
    parser.add_argument("--internal-orange", type=float, default=DEFAULT_HIGHLIGHT_THRESHOLDS.internal_orange, help="Internal-standard percent-error threshold for orange; higher values use dark orange.")
    parser.add_argument("--blank-light-orange", type=float, default=DEFAULT_HIGHLIGHT_THRESHOLDS.blank_light_orange, help="Blank-sample ppb tolerance before orange highlighting.")
    parser.add_argument("--blank-orange", type=float, default=DEFAULT_HIGHLIGHT_THRESHOLDS.blank_orange, help="Blank-sample ppb threshold for orange; higher values use dark orange.")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: Input file does not exist: {args.input}")
        return 1

    try:
        calibration_range = CalibrationRange(
            orange_min=args.orange_min,
            green_min=args.green_min,
            green_max=args.green_max,
        )
        highlight_thresholds = HighlightThresholds(
            internal_light_orange=args.internal_light_orange,
            internal_orange=args.internal_orange,
            blank_light_orange=args.blank_light_orange,
            blank_orange=args.blank_orange,
        )
        return process_icp_file(
            args.input,
            args.output,
            parse_sheet_arg(args.sheet),
            normalize_sample_initials(args.initials),
            calibration_range,
            highlight_thresholds,
        )
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
